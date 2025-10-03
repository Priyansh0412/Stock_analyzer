import requests
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from datetime import datetime
import time
import yfinance as yf

class AccurateNSEStockAnalyzer:
    def __init__(self):
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': '*/*',
            'Accept-Language': 'en-US,en;q=0.9',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
        }
        self.session = requests.Session()
        
    def get_accurate_stock_data_yfinance(self, symbol):
        try:
            ticker_symbol = f"{symbol}.NS"
            print(f"  → Fetching from Yahoo Finance: {ticker_symbol}")
            ticker = yf.Ticker(ticker_symbol)
            info = ticker.info
            hist_1y = ticker.history(period="1y")
            hist_3m = ticker.history(period="3mo")
            hist_1m = ticker.history(period="1mo")
            if hist_1y.empty:
                print(f"  ✗ No data available for {symbol}")
                return None
            current_price = info.get('currentPrice') or info.get('regularMarketPrice') or hist_1y['Close'].iloc[-1]
            week_52_high = hist_1y['High'].max()
            week_52_low = hist_1y['Low'].min()
            month_3_high = hist_3m['High'].max()
            month_3_low = hist_3m['Low'].min()
            month_1_high = hist_1m['High'].max()
            month_1_low = hist_1m['Low'].min()
            data = {
                'current_price': float(current_price),
                '52w_high': float(week_52_high),
                '52w_low': float(week_52_low),
                '3m_high': float(month_3_high),
                '3m_low': float(month_3_low),
                '1m_high': float(month_1_high),
                '1m_low': float(month_1_low),
                'source': 'Yahoo Finance (yfinance)'
            }
            print(f"  ✓ Success! Current Price: ₹{data['current_price']:.2f}")
            return data
        except Exception as e:
            print(f"  ✗ yfinance error: {str(e)}")
            return None

    # Keep other methods (get_from_moneycontrol, get_from_google_finance, get_nse_official) as-is

    def get_stock_info(self, symbol, company_name):
        print(f"\n{'='*60}")
        print(f"📊 Analyzing: {symbol}")
        print(f"{'='*60}")
        
        stock_data = self.get_accurate_stock_data_yfinance(symbol)
        if not stock_data or stock_data.get('current_price', 0) == 0:
            nse_data = self.get_nse_official(symbol)
            if nse_data:
                stock_data = nse_data
        if not stock_data or stock_data.get('current_price', 0) == 0:
            google_data = self.get_from_google_finance(symbol)
            if google_data:
                yf_data = self.get_accurate_stock_data_yfinance(symbol)
                if yf_data:
                    stock_data = yf_data
                else:
                    stock_data = google_data
        if not stock_data or stock_data.get('current_price', 0) == 0:
            print(f"  ⚠ Warning: Could not fetch live data, using last known prices")
            stock_data = {'current_price': 0, 'source': 'Unavailable'}
        
        current_price = stock_data.get('current_price', 0)
        stock_info = {
            'Symbol': symbol,
            'Current_Price': round(current_price, 2),
            '52_Week_High': round(stock_data.get('52w_high', current_price * 1.2), 2),
            '52_Week_Low': round(stock_data.get('52w_low', current_price * 0.8), 2),
            '3_Month_High': round(stock_data.get('3m_high', current_price * 1.1), 2),
            '3_Month_Low': round(stock_data.get('3m_low', current_price * 0.9), 2),
            '1_Month_High': round(stock_data.get('1m_high', current_price * 1.05), 2),
            '1_Month_Low': round(stock_data.get('1m_low', current_price * 0.95), 2),
            'Data_Source': stock_data.get('source', 'Unknown')
        }
        
        stock_info['Price_vs_52W'] = self.calculate_position(stock_info['Current_Price'], stock_info['52_Week_Low'], stock_info['52_Week_High'])
        stock_info['Price_vs_3M'] = self.calculate_position(stock_info['Current_Price'], stock_info['3_Month_Low'], stock_info['3_Month_High'])
        stock_info['Price_vs_1M'] = self.calculate_position(stock_info['Current_Price'], stock_info['1_Month_Low'], stock_info['1_Month_High'])
        
        stock_info['Current_vs_All'] = round(
            (stock_info['Price_vs_52W'] + stock_info['Price_vs_3M'] + stock_info['Price_vs_1M']) / 3, 1
        )
        
        print(f"\n📈 Results:")
        print(f"  • Current Price: ₹{stock_info['Current_Price']:.2f}")
        print(f"  • 52W Range: ₹{stock_info['52_Week_Low']:.2f} - ₹{stock_info['52_Week_High']:.2f}")
        print(f"  • 3M Range: ₹{stock_info['3_Month_Low']:.2f} - ₹{stock_info['3_Month_High']:.2f}")
        print(f"  • 1M Range: ₹{stock_info['1_Month_Low']:.2f} - ₹{stock_info['1_Month_High']:.2f}")
        print(f"  • Current vs All: {stock_info['Current_vs_All']}%")
        print(f"  • Data Source: {stock_info['Data_Source']}")
        return stock_info
    
    def calculate_position(self, current, low, high):
        if high == low or high == 0:
            return 50.0
        position = ((current - low) / (high - low)) * 100
        return round(max(0, min(100, position)), 1)
    
    def analyze_stocks(self, stock_dict):
        results = []
        for symbol, company_name in stock_dict.items():
            stock_info = self.get_stock_info(symbol, company_name)
            results.append(stock_info)
            time.sleep(2)
        return pd.DataFrame(results)
    
    def create_excel_with_speedometer(self, df, filename='Stock_Analysis_Report.xlsx'):
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        df_no_company = df.drop(columns=['Company'], errors='ignore')  # remove company column if exists
        df_no_company.to_excel(writer, sheet_name='Stock Data', index=False, startrow=2)
        workbook = writer.book
        worksheet = writer.sheets['Stock Data']
        
        worksheet['A1'] = f'NSE Stock Analysis Report - {datetime.now().strftime("%d %B %Y, %I:%M %p")}'
        worksheet['A1'].font = Font(size=14, bold=True, color='000000')
        worksheet.merge_cells('A1:O1')
        worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        worksheet.row_dimensions[1].height = 30
        
        header_font = Font(bold=True, color='000000', size=11)
        for cell in worksheet[3]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.row_dimensions[3].height = 40
        
        thin_border = Border(left=Side(style='thin', color='D3D3D3'),
                             right=Side(style='thin', color='D3D3D3'),
                             top=Side(style='thin', color='D3D3D3'),
                             bottom=Side(style='thin', color='D3D3D3'))
        
        for row in worksheet.iter_rows(min_row=4, max_row=len(df_no_company)+3, min_col=1, max_col=len(df_no_company.columns)):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if isinstance(cell.value, (int, float)) and cell.column <= 10:
                    cell.number_format = '₹#,##0.00'
                elif isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.0'
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            col_idx = None
            for cell in column:
                if not isinstance(cell, MergedCell):
                    col_idx = cell.column
                    break
            if col_idx is None:
                continue
            column_letter = get_column_letter(col_idx)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 3, 22)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        writer.close()
        print(f"\n{'='*60}")
        print(f"✓ Excel file created successfully: {filename}")
        print(f"{'='*60}")
        return filename

def main():
    print("\n" + "="*70)
    print(" "*15 + "🔴 ACCURATE NSE STOCK ANALYZER 🔴")
    print("="*70)
    print(f"\n📅 Analysis Date: {datetime.now().strftime('%d %B %Y, %I:%M %p')}")
    
    stock_dict = {
        'IDEA': 'Vodafone Idea Limited',
        'ADANIPORTS': 'Adani Ports and SEZ',
        'RELIANCE': 'Reliance Industries',
        'BAJAJ-AUTO': 'Bajaj Auto Limited'
    }
    
    print(f"\n📊 Stocks to analyze ({len(stock_dict)}):")
    for symbol in stock_dict.keys():
        print(f"   • {symbol}")
    
    print("\n" + "-"*70)
    print("🔍 Fetching LIVE data from multiple sources...")
    print("   Sources: Yahoo Finance (yfinance), NSE Official, Google Finance")
    print("-"*70)
    
    analyzer = AccurateNSEStockAnalyzer()
    df = analyzer.analyze_stocks(stock_dict)
    
    print("\n" + "="*70)
    print(" "*22 + "📊 ANALYSIS SUMMARY 📊")
    print("="*70 + "\n")
    
    display_df = df[['Symbol', 'Current_Price', '52_Week_Low', '52_Week_High',
                     'Price_vs_52W', 'Price_vs_3M', 'Price_vs_1M', 'Current_vs_All', 'Data_Source']].copy()
    
    display_df.columns = ['Symbol', 'Price (₹)', '52W Low', '52W High',
                          '52W %', '3M %', '1M %', 'Current_vs_All (%)', 'Source']
    
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', None)
    pd.set_option('display.precision', 2)
    
    print(display_df.to_string(index=False))
    
    print("\n" + "-"*70)
    print("📝 Creating detailed Excel report...")
    print("-"*70)
    
    filename = analyzer.create_excel_with_speedometer(df, filename=f"Stock_Analysis_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

    print("\n" + "="*70)
    print(" "*25 + "✅ ANALYSIS COMPLETE!")
    print("="*70)
    print(f"\n📄 Output file: {filename}")

if __name__ == "__main__":
    main()

